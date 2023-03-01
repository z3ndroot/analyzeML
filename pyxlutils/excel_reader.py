from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, filename):
        self.filename = filename
        self._wb = None
        self._names_columns_dict = None

    @property
    def wb(self):
        if self._wb is None:
            self._wb = load_workbook(filename=self.filename)
        return self._wb

    @property
    def sheet_names(self):
        return self.wb.sheetnames

    def columns_names(self, name_sheet):
        worksheet = self.wb[name_sheet]
        iter_cols = worksheet.iter_cols(values_only=True)
        columns_names = [title[0] for title in iter_cols if title[0] is not None]

        self._names_columns_dict = {title: num for num, title in enumerate(
            columns_names,
        )}

        return columns_names

    def get_messages(self, sheet_name, column_name):
        sheet = self.wb[sheet_name]
        index_column = self._names_columns_dict[column_name]
        column_values = [row[index_column] for row in sheet.iter_rows(min_row=2, values_only=True)]

        return column_values
