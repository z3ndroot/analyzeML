from openpyxl import load_workbook
from pypred import Predicate


class FileXSLX:
    def __init__(self, filename):
        self.filename = filename
        self._wb = None
        self._names_columns_dict = None

    @property
    def wb(self):
        if self._wb is None:
            self._wb = load_workbook(filename=self.filename)
        return self._wb

    @property
    def sheet_names(self):
        return self.wb.sheetnames

    def columns_names(self, name_sheet):
        worksheet = self.wb[name_sheet]
        iter_cols = worksheet.iter_cols(
            max_row=1,
            max_col=worksheet.max_column,
            values_only=True
        )
        columns_names = [title[0] for title in iter_cols if title[0] is not None]

        self._names_columns_dict = {title: num for num, title in enumerate(
            columns_names,
            start=1,
        )}

        return columns_names

    # TODO...
